import openpyxl
from pathlib import Path
from typing import Any, Dict, List

def read_excel_rows(xlsx_path: str, sheet_name: str) -> List[Dict[str, Any]]:
    """Read the Excel sheet starting from row 2 and return rows keyed by headers."""
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")
    if path.suffix.lower() not in {".xlsx", ".xls"}:
        raise ValueError(f"Excel file must be .xlsx or .xls, got: {path.suffix}")

    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {xlsx_path}")
        sheet = workbook[sheet_name]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell).strip() if cell is not None else f"column_{idx + 1}"
                   for idx, cell in enumerate(header_row)]

        data: List[Dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(cell is not None for cell in row):
                continue

            row_data = {
                headers[idx] if idx < len(headers) else f"column_{idx + 1}": cell
                for idx, cell in enumerate(row)
            }
            data.append(row_data)

        return data
    finally:
        workbook.close()


def read_data_by_columns(xlsx_path: str, sheet_name: str, column_names: List[str]) -> List[Dict[str, Any]]:
    """Read only the requested columns from the sheet, starting from row 2."""
    if not isinstance(column_names, list) or not column_names:
        raise ValueError("column_names must be a non-empty list")
    rows = read_excel_rows(xlsx_path, sheet_name)
    return [{col: row.get(col) for col in column_names} for row in rows]


def read_project_name_from_row(xlsx_path: str, row_number: int = 2, sheet_name: str = "ProjectName") -> str:
    """Read the project name from a specific row in the ProjectName sheet."""
    if row_number < 2:
        raise ValueError("row_number must be 1 or greater")

    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")
    if path.suffix.lower() not in {".xlsx", ".xls"}:
        raise ValueError(f"Excel file must be .xlsx or .xls, got: {path.suffix}")

    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {xlsx_path}")

        sheet = workbook[sheet_name]
        project_name = sheet.cell(row=row_number, column=1).value
        if project_name is None:
            raise ValueError(f"No project name found in row {row_number}, column 1 of sheet '{sheet_name}'")

        return str(project_name).strip()
    finally:
        workbook.close()


def read_project_names(xlsx_path: str, sheet_name: str = "ProjectName") -> List[str]:
    """Read the ProjectName sheet and return project name values."""
    rows = read_excel_rows(xlsx_path, sheet_name)
    if not rows:
        return []

    first_column = next(iter(rows[0].keys()))
    return [row.get(first_column) for row in rows if row.get(first_column) is not None]


if __name__ == "__main__":
    path = Path(__file__).resolve().parent.parent / "test_data" / "test_automation_data.xlsx"

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    rows = read_excel_rows(str(path), "LoginTests")
    print(f"Read {len(rows)} rows from {path.name}")
    for index, row in enumerate(rows, start=2):
        print(f"Row {index}: {row}")


